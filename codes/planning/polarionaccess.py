import polarion.polarion as polarion
import numpy as np
import csv
from openpyxl import load_workbook, Workbook
from openpyxl.styles import PatternFill,Font,Alignment
from datetime import datetime

username,password=np.loadtxt("secret.txt",dtype="str",unpack=1,delimiter=",") # comma separated file with username and password
#print(username,password)
client = polarion.Polarion('https://polarion.astron.nl/polarion', username, password)
project = client.getProject('METIS')
#workitem135 = project.getWorkitem('METIS-135')

set = project.searchWorkitem('type:(system_test_case) AND NOT status:discarded')

wi = project.getWorkitem('METIS-10970')
getkeys = wi.getAllowedCustomKeys() 
print(getkeys) # ['testSteps', 'tccategory', 'subsys', 'prereq', 'hazard', 'ait-id', 'owner', 'maturity', 'pip_recipes']
excel_id = []
aitid = []
pip = []

all_aitid = []
all_pip = []
all_subsys=[]
all_hazard=[]
excel_description=[]
excel_owner = []
excel_subsys=[]
excel_hazard=[]
excel_aitid=[]
excel_pip=[]
excel_status=[]
for sss in set: #set[0:5]: # artifically limit it to 5 entries
    ci = project.getWorkitem(sss['id'])
    #print()
    #help(ci)
    excel_id.append(sss['id'])
    excel_description.append(ci.title)
    excel_status.append(ci.status.id)
    try:
        pass
        #print(ci.status.id) # use this to try out keywords
    except:
        print("didn't work")
    caitid = ci.getCustomField('ait-id')
    csubsys = ci.getCustomField('subsys')
    chazard = ci.getCustomField('hazard')
    cowner = ci.getCustomField('owner').id
    cpip = ci.getCustomField('pip_recipes')
    pip.append(cpip)
    aitid.append(caitid)
    excel_owner.append(cowner)
    if type(caitid) != type(None):
        temp_aitid=[]
        for iii in caitid.EnumOptionId:
            all_aitid.append(iii.id)
            temp_aitid.append(iii.id)
        temp_aitid.sort()
        excel_aitid.append(",".join(temp_aitid))
    else:
        print(ci.id+" has no AIT-ID")
        excel_aitid.append("Not set")
    if type(cpip) != type(None):
        temp_pip=[]
        for iii in cpip.EnumOptionId:
            all_pip.append(iii.id)
            temp_pip.append(iii.id)
        temp_pip.sort()
        excel_pip.append(",".join(temp_pip))
    else:
        print(ci.id+" has no associated PIP recipes")
        excel_pip.append("Not set")
    if type(csubsys) != type(None):
        temp_subsys=[]
        for iii in csubsys.EnumOptionId:
            all_subsys.append(iii.id)
            temp_subsys.append(iii.id)
        temp_subsys.sort()
        excel_subsys.append(",".join(temp_subsys))
    else:
        print(ci.id+" has no associated subsystems")
        excel_subsys.append("Not set")
    if type(chazard) != type(None):
        temp_hazard=[]
        for iii in chazard.EnumOptionId:
            all_hazard.append(iii.id)
            temp_hazard.append(iii.id)
        temp_hazard.sort()
        excel_hazard.append(",".join(temp_hazard))
    else:
        print(ci.id+" has no associated hazards")
        excel_hazard.append("Not set")

# remko stuff \/\/ used for AIV phase - pip recipe matching matrix

uaitid = np.unique(all_aitid)
upip = np.unique(all_pip)

naitid = len(uaitid)
npip = len(upip)
kaitid = {}
for xxx,iii in zip(range(naitid), uaitid):
    kaitid[iii] = xxx
kpip = {}
for xxx,iii in zip(range(npip), upip):
    kpip[iii] = xxx
matcharr = np.zeros((naitid, npip))

for aaa, ppp in zip(aitid, pip):
    if (aaa != None) & (ppp != None):
        for iii in aaa.EnumOptionId:
            for jjj in ppp.EnumOptionId:
                matcharr[kaitid[iii.id],kpip[jjj.id]] = 1

with open('aitid.csv', 'w', newline='') as fp:
    writer = csv.writer(fp)
    writer.writerow(uaitid)
with open('pip.csv', 'w', newline='') as fp:
    writer = csv.writer(fp)
    writer.writerow(upip)
np.savetxt('match.csv', np.transpose(matcharr), delimiter=",")

# making Excel file \/\/\/
wb=Workbook()
w0=wb.active
w0.title="Title page"



current_dateTime = datetime.utcnow()



w0.append(["Authors:",'Remko Stuik, Gilles Otten, Adrian Glauser'])
w0.append(["Parsing date (UTC):",current_dateTime.strftime('%Y %B %d - %H:%M:%S')])

ws=wb.create_sheet()
ws.title="System test cases"
wslist=[]

#determine sortorder
temparr=[]
for iii in excel_id:
    temparr.append(np.int32(iii.split("-")[1])) # throw away "METIS-" and keep number XXXX from METIS-XXXX
sortnum=np.argsort(temparr)[::-1] # sort large to small

#make list to be written
wslist.append(excel_id)
wslist.append(excel_description)
wslist.append(excel_subsys)
wslist.append(excel_owner)
wslist.append(excel_hazard)
wslist.append(excel_aitid)
wslist.append(excel_status)
wslist.append(excel_pip)
wsarr=np.array(wslist).T[sortnum]
wslist=wsarr.tolist() # sort list based on polarion ID as a ndarray, transpose rows and columns and convert back to list
print(wslist)

headernames=["POLARION ID","Description","Subsystem","Owner","Hazard","AIT IDs","Status","PIP recipes"]
ws.append(headernames)
# write to file
for row in wslist:
   ws.append(row)
   
#ws.insert_rows(1)
#ws['A1']="POLARION ID"
#ws['B1']="Description"
#ws['C1']='Subsystem'
#ws['D1']="Owner"
#ws['E1']="Hazard"
#ws['F1']="AIT IDs"
#ws['G1']="Status"
#ws['H1']="PIP recipes"

# make new worksheet with recipe matrix
wb.create_sheet("AIV-Recipe matrix")
ws_mat=wb['AIV-Recipe matrix']
# insert matrix
ws_mat.append(uaitid.tolist())
for row in matcharr.T.tolist():
   ws_mat.append(row)
ws_mat.insert_cols(1) # insert empty column at start for PIP recipe names

dim=np.shape(matcharr)
for j in np.arange(dim[0]):
    c=ws_mat.cell(row=1,column=j+2)
    c.alignment = Alignment(horizontal='center')
for i in np.arange(dim[1]):
    for j in np.arange(dim[0]):
        c=ws_mat.cell(row=i+2,column=j+2)
        if c.value == 1:
            ws_mat.cell(row=i+2,column=j+2).fill= PatternFill(start_color="a8d08d", end_color="a8d08d", fill_type = "solid")
            c.font=Font(color="a8d08d") # make green text
            #ws_mat.cell(row=i+2,column=j+2).value=""
        elif c.value == 0:
            c.font=Font(color="FFFFFF") # make white text
        else:
            print("enumerated empty cell?")
        
for i,up in enumerate(upip): # add names in first column
    ws_mat.cell(row=2+i,column=1).value=up
for i in np.arange(len(headernames)+1): # make boldface top row
    ws.cell(row=1,column=i+1).font = Font(bold=True)

# from sheet 1 enumerate all cells and check if they are "Not set" 
dim=np.shape(wsarr)
for i in np.arange(dim[0]):
    for j in np.arange(dim[1]):
        c=ws.cell(row=i+2,column=j+2)
        if c.value in ["Not set","N/A or not set"]:
            ws.cell(row=i+2,column=j+2).fill= PatternFill(start_color="ff0000", end_color="ff0000", fill_type = "solid")
        elif c.value == "na":
            c.value="N/A"
        else:
            pass
    
def resizecol(wsheet,size=4): ## automatically adjust column widths of sheet to be readable, minimal width = size (default=4) characters
    for column_cells in wsheet.columns: 
        length = max(len(str(cell.value)) for cell in column_cells)
        wsheet.column_dimensions[column_cells[0].column_letter].width = np.max([length*1.05,size])    
resizecol(w0)
resizecol(ws)
resizecol(ws_mat)

# freeze panes
ws.freeze_panes=ws['C2']
ws_mat.freeze_panes=ws['B2']
# save workbook as excel file
wb.save("metis_systemtestcase_python_"+current_dateTime.strftime("%Y%m%d")+".xlsx")
wb.close